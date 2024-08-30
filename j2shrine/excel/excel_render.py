from dataclasses import dataclass
import openpyxl
from ..loader import Loader
from ..context import RenderContext
from ..processors import Processor

@dataclass
class CellPosition:
    """セル位置を表す行番号と列名クラス"""
    row:int
    col:str


class ExcelLoader(Loader):

    # jinja2テンプレートの生成
    def __init__(self, *, context: RenderContext, processor: Processor):
        super().__init__(context=context, processor=processor)
        # カラム名は追加される可能性があるためcopyする
        self.cols = context.names.copy() if context.names is not None else []

    def loading(self):
        """Workbookの中身をcontextに従って読み取る"""
        # Workbookを開く
        reader = openpyxl.load_workbook(self.context.source, data_only=True)
        # シート内セルの読み込み範囲
        (start, end) = self.context.read_range
        # 読み込むシートの範囲
        (first_sheet, end_sheet) = self.context.sheets

        # 読み取りシート範囲の指定
        if (end_sheet is None):
            # endがNoneの場合最後のシートまで読む
            end_sheet = len(reader.worksheets)-1

        results = []
        sheet_idx = first_sheet
        while sheet_idx <= end_sheet:
            sheet = reader.worksheets[sheet_idx]

            # シート読込み開始
            sheet_data = {
                # シート名
                'name': reader.sheetnames[sheet_idx], 
                # 絶対座標指定セル
                'abs': self.absolute_cells(sheet=sheet, cells=self.context.absolute),
                # シート内容の行
                'rows': [],
            }
            # シート内の指定範囲のセルの読み込み
            for line_no, row in enumerate(sheet.iter_rows(min_col=start.col, min_row=start.row,
                                       max_col=end.col, max_row=end.row, )):
                # rowsを一行ずつ処理
                sheet_data['rows'].append(self.read_row(line_no=line_no, cells=row))

            results.append(sheet_data)
            sheet_idx = 1 + sheet_idx
        return {
            'sheets': results,
            'params': self.context.parameters
        }

    def read_row(self, *, line_no:int, cells:list):
        """カラムのlistをdictに変換する。dictのキーはcontext.namesで与えられた名前か、連番になる。"""
        line = {}
        for index, cell in enumerate(cells):
            name = self.get_name(index)
            # カラムから値を取り出す
            line[name] = self.read_value(name=name, cell=cell)
        return line

    def absolute_cells(self, *, sheet, cells: dict):
        """絶対座標セル値の取得"""
        cell_values = {}
        for name, addr in cells.items():
            cell_values[name] = self.read_value(name=name, cell=sheet[addr])
        return cell_values

    def read_value(self, *, name, cell):
        """value属性が存在すればセルの値を取得する"""
        return cell.value if hasattr(cell, 'value') else None
