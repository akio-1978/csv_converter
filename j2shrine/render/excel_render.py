import openpyxl
from . base_render import Render, RenderContext
from . excel_custom_filter import excel_time

# 取得するシート a a-b a-
# 取得するカラムのレンジ A A-Z A-
# 取得する行 a a-b a-
# option
# ヘッダ行
# ヘッダ直接指定
# 追加取得セル...
class ExcelRenderContext(RenderContext):
    def __init__(self, *, template=None, template_encoding='utf8', parameters={}):
        super().__init__(template=template, template_encoding=template_encoding, parameters=parameters)
        self.encoding = 'utf8'
        self.sheets = '1'
        self.header_prefix='col_'
        self.header_row = None     # ex. 1
        self.read_range = None # A2:C4  (read cells from A2 to C4) or A2:C (read cells in all rows from 2)
        self.limit = None
        self.extra =[]

class  ReadSetting:
    def __init__(self, *, sheet_left:int, sheet_right:int, left_row:str, left_column:str, right_row:str, right_column:str) -> None:
        self.sheet_left = sheet_left
        self.sheet_right = sheet_right
        self.left_row = left_row
        self.left_column = left_column
        self.right_row = right_row
        self.right_column = right_column

    def get_sheet_right(self, sheets):
        if self.sheet_right is not None:
            return self.sheet_right
        else:
            return len(sheets.worksheets) -1

class ExcelRender(Render):

    ALPHABET_TABLE = '0ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    TO_END = -1

    # jinja2テンプレートの生成
    def __init__(self, *, context :ExcelRenderContext):
        super().__init__(context = context)
        # シートからの取得範囲は最初に特定する
        self.setup_range()

    def install_filters(self, *, environment):
        super().install_filters(environment=environment)
        environment.filters['excel_time'] = excel_time

    def setup_range(self):

        (left_row, left_column, right_row, right_column) = self.get_read_range(self.context.read_range)
        (sheet_left, sheet_right) = self.get_sheet_range(self.context.sheets)

        return ReadSetting(sheet_left=sheet_left, sheet_right=sheet_right,
            left_row=left_row, left_column=left_column, right_row=right_row, right_column=right_column)

    def build_reader(self, *, source :any) :
        # 既にブックで渡された場合、そのまま返す
        if ( isinstance(source, openpyxl.Workbook)):
            return source
        # ファイルの場合はロードする
        return openpyxl.load_workbook(source, data_only=True)

    def read_source(self, *, reader):

        setting = self.setup_range()
        sheet_right = setting.get_sheet_right(sheets=reader.worksheets)

        all_sheets = []
        sheet_idx = setting.sheet_left
        while sheet_idx <= sheet_right:
            sheet = reader.worksheets[sheet_idx]
            # ヘッダの読込み
            headers = self.read_headers(sheet=sheet, setting=setting)

            # コンテンツ読込み
            sheet_data = {
                'name' : reader.sheetnames[sheet_idx],
                'rows' : [],
                'headers' : headers,
                'extra' : self.read_extra_cells(sheet= sheet)
            }
            for row in sheet.iter_rows(min_col=setting.left_column, min_row=setting.left_row , 
                    max_col=setting.right_column, max_row=setting.right_row, ):
                sheet_data['rows'].append(self.columns_to_dict(columns = row, headers = headers))

            all_sheets.append(sheet_data)
            sheet_idx = 1 + sheet_idx
        return all_sheets

    def read_extra_cells(self, *, sheet):
        cells = {}
        for addr in self.context.extra:
            cells[addr] = sheet[addr].value
        return cells

    def finish(self, *, result):

        final_result = {
            'sheets' : result,
            'params' : self.context.parameters
        }

        return final_result


    def read_headers(self, *, sheet:openpyxl.worksheet.worksheet, setting:ReadSetting):
        headers = []

        # どちらを通っているかはっきりしていない
        # ヘッダはシートごとの独立させるか否かを検討する

        if self.context.header_row is not None:
            for row in sheet.iter_rows(min_col=setting.left_column, min_row=int(self.context.header_row),
                                        max_col=setting.right_column, max_row=int(self.context.header_row)):
                for cell in row:
                    headers.append(cell.value)
        else: 
            for row in sheet.iter_rows(min_col=setting.left_column, min_row=1,
                                        max_col=setting.right_column, max_row=1):
                for cell in row:
                    headers.append(cell.column_letter)

        return headers

    # カラムのlistをdictに変換する。dictのキーはself.headers
    def columns_to_dict(self, *, columns, headers):
        line = {}
        # カラムとヘッダの長さは揃っていることが前提
        for header, column in zip(headers, columns):
            # カラム単体の変換処理を行う
            line[header] = self.read_column(name = header, column = column)
        return line

    def columns_dict(self, *, columns_dict):
        return columns_dict

    # hook by every column
    def read_column(self, *, name, column):
        # データの取り出し
        return column.value
    
    def column_number(self, *, column:str):
        fulldigit = column.zfill(3)

        number = 0
        number = number + (int( ExcelRender.ALPHABET_TABLE.find(fulldigit[0])) * 26 * 26)
        number = number + (int(ExcelRender.ALPHABET_TABLE.find(fulldigit[1])) * 26)
        number = number + (int(ExcelRender.ALPHABET_TABLE.find(fulldigit[2])))

        return number

    def get_sheet_range(self, *, sheets_range:str):

        (left, to, right) = sheets_range.partition(':')

        if to == ':':
            if right == '':
                return (int(left) -1, None)
            else:
                return (int(left) -1, int(right) -1)

        return (int(left) -1, int(left) -1)

    def get_read_range(self, *, arg_range:str, all=None):

        (left, delim, right) = arg_range.partition(':')

        if left == arg_range:
            raise ValueError('invalid range: ' + arg_range)

        (left_row, left_column) = self.get_coordinate(left)
        (right_row, right_column) = self.get_coordinate(right)

        return (left_row, left_column, right_row, right_column)

    def get_coordinate(self, *, coordinate:str):
        if coordinate.isalpha() and coordinate.isascii():
            # args is column only. read all rows.
            column = openpyxl.utils.cell.column_index_from_string(coordinate)
            return (None, column)
        else:
            # full specified coordinate.
            return openpyxl.utils.cell.coordinate_to_tuple(coordinate)


    def get_cell_value(self, *, sheet, cell):
        return sheet[cell].value

